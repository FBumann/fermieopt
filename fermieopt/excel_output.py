# -*- coding: utf-8 -*-
import os
from typing import Union, List, Literal

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference,LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt

from fermieopt.flixPostXL import flixPostXL

### Utility functions
def resample_data(data_frame: Union[pd.DataFrame, np.ndarray], target_years: List[int], resampling_by: Literal["YE", "d", "h"],
                   resampling_method: Literal["sum", "mean", "min","max"], initial_sampling_rate: str = "h") -> pd.DataFrame:
    '''
    Parameters
    ----------
    data_frame : Union[pd.DataFrame, np.ndarray]
        DataFrame or array containing data. Number of rows must match the initial sampling rate (safety check):
        8760 ("h") (default) or 365 ("d") per year.
    target_years : List[int]
        Target years for the new index of the DataFrame
    resampling_by : str
        "h" for hourly resampling
        "d" for daily resampling
        "YE" for yearly resampling
    resampling_method : str
        "mean" for mean value
        "sum" for sum value
        "max" for max value
        "min" for min value
    initial_sampling_rate : str
        "h" for hourly data (8760 values per year)
        "d" for daily data (365 values per year)

    Returns
    -------
    pd.DataFrame
    '''
    df = pd.DataFrame(data_frame)
    df.index = range(len(df))  # reset index

    if len(df)/8760 == len(target_years) and initial_sampling_rate == "h":
        length_per_year = 8760
    elif len(df)/365 == len(target_years) and initial_sampling_rate == "d":
        length_per_year = 365
    else:
        raise ValueError("length of dataframe and initial_sampling_rate must match: "
                         "8760 rows/year ('H') or 365 rows/year 'D'.")

    if not isinstance(target_years, list):
        target_years = [target_years]

    # create new TS for resampling, without the 29. February (filtering leap years)
    for i, year in enumerate(target_years):
        dt = pd.date_range(start=f'1/1/{year}', end=f'01/01/{year + 1}', freq=initial_sampling_rate)[:-1]
        dt = dt[~((dt.month == 2) & (dt.day == 29))]  # Remove leap year days
        df.loc[i * length_per_year:(i + 1) * length_per_year - 1, 'Timestamp'] = dt
    df = df.set_index('Timestamp')

    if resampling_method == "sum":
        df = df.resample(resampling_by).sum()
    elif resampling_method == "mean":
        df = df.resample(resampling_by).mean()
    elif resampling_method == "min":
        df = df.resample(resampling_by).min()
    elif resampling_method == "max":
        df = df.resample(resampling_by).max()
    else:
        raise ValueError("Invalid resampling method")

    # Drop all rows that aren't in the years specified in target_years
    lst = [row for row in df.index if row.year not in target_years]
    df = df.drop(index=lst)

    df = df.loc[~((df.index.month == 2) & (df.index.day == 29))]  # Remove leap year days again

    if resampling_by == "YE":
        df = df.set_index(df.index.year)  # setting the index to the plain year. No datetime anymore

    return df

def rs_in_two_steps(data_frame: Union[pd.DataFrame, np.ndarray], target_years: List[int], resampling_by: Literal["d", "YE"],
                    initial_sampling_rate: str = "h") -> pd.DataFrame:
    '''
    Parameters
    ----------
    data_frame : Union[pd.DataFrame, np.ndarray]
        DataFrame or array containing data. Number of rows must match the initial sampling rate (safety check):
        8760 ("h") (default) or 365 ("d") per year.
    target_years : List[int]
        Years for resampling
    resampling_by : str
        "d" for daily resampling
        "YE" for yearly resampling
    initial_sampling_rate : str
        "h" for hourly data
        "d" for daily data
    Returns
    -------
    pd.DataFrame
        Resampled DataFrame with new columns:
        ["Tagesmittel", "Minimum (Stunde)", "Maximum (Stunde)"]
        or new Columns:
        ["Jahresmittel", "Minimum (Tagesmittel)", "Maximum (Tagesmittel)"],
        depending on chosen "resampling_by"
    '''

    # Determine base resampling method and new columns based on resampling_by
    if resampling_by == "d":
        rs_method_base = "h"
        new_columns = ["Tagesmittel", "Minimum (Stunde)", "Maximum (Stunde)"]
    elif resampling_by == "YE":
        rs_method_base = "d"
        new_columns = ["Jahresmittel", "Minimum (Tagesmittel)", "Maximum (Tagesmittel)"]
    else:
        raise ValueError("Invalid value for resampling_by. Use 'D' for daily or 'Y' for yearly.")


    # Base resampling
    df_resampled_base = resample_data(data_frame, target_years, rs_method_base, "mean", initial_sampling_rate)

    # Resample for min, max, and mean
    min_y = resample_data(df_resampled_base, target_years, resampling_by, "min", rs_method_base)
    max_y = resample_data(df_resampled_base, target_years, resampling_by, "max", rs_method_base)
    mean_y = resample_data(df_resampled_base, target_years, resampling_by, "mean", rs_method_base)

    # Concatenate results
    df_result = pd.concat([mean_y, min_y, max_y], axis=1)
    df_result.columns = new_columns

    return df_result

def reorder_columns(df:pd.DataFrame, not_sorted_columns: List[str] = None):
    '''
    Order a DataFrame by a custom function, excluding specified columns from sorting, and setting them as the first columns.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame.
    not_sorted_columns : List[str], optional
        Columns to exclude from sorting and set as the first columns, by default None.

    Returns
    -------
    pd.DataFrame
        DataFrame with the desired column order.
    '''
    if isinstance(df, pd.Series): df = df.to_frame().T
    sorted_columns = sorted(df.columns, key=lambda x: x.lower())
    sorted_df = df.reindex(columns=sorted_columns)

    # Select the remaining columns excluding the first two
    if not_sorted_columns is None:
        other_columns = [col for col in sorted_df.columns]
        # Create a new DataFrame with the desired column order
        new_order_df = sorted_df[other_columns]
    else:
        other_columns = [col for col in sorted_df.columns if col not in not_sorted_columns]

        # Create a new DataFrame with the desired column order
        new_order_df = pd.concat([sorted_df[not_sorted_columns], sorted_df[other_columns]], axis=1)

    return new_order_df

def df_to_excel_w_chart(df: pd.DataFrame, filepath: str, title: str, ylabel: str, xlabel: str, style:Literal["bar","line"]="bar", bar_style:Literal["stacked","clustered"] = "stacked"):
    """
    Write DataFrame to an Excel file with a stacked bar chart.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing the data to be written.
    filepath : str
        The path to the Excel file. If the file doesn't exist, a new one will be created.
    title : str
        The title of the sheet and chart.
    ylabel : str
        The label for the y-axis of the chart.
    xlabel : str
        The label for the x-axis of the chart.

    Returns
    -------
    None

    Notes
    -----
    This function writes the provided DataFrame to an Excel file and adds a stacked bar chart to a new sheet in the workbook.
    If the sheet with the given title already exists, it is removed before adding the new sheet.
    The stacked bar chart is created based on the DataFrame structure, with columns as categories and rows as data points.
    The chart is positioned at cell "D4" in the sheet.

    """
    try:
        wb = load_workbook(filepath)
    except:
        template_path = Path(__file__).resolve().parent / "resources" / "Template_blanco.xlsx"

        wb = load_workbook(template_path)

    # Check if the sheet already exists
    if title in wb.sheetnames:
        sheet = wb[title]
        wb.remove(sheet)

    # Add the sheet to the workbook
    sheet = wb.create_sheet(title)

    # Remove the index and save it as a column
    df = df.reset_index()
    # Write the data starting from the second row
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # Create the Data and References
    data = Reference(sheet, min_col=2, min_row=1, max_col=df.shape[1], max_row=df.shape[0] + 1)
    labels = Reference(sheet, min_col=1, min_row=2, max_row=df.shape[0] + 1)

    # Create a stacked bar chart
    if style=="bar":
        chart = BarChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        # Stacked bar plot
        chart.type = "col"
        chart.grouping = bar_style
        if bar_style == "stacked":
            chart.overlap = 100
            chart.gapWidth = 0  # Adjust the gap between bars (e.g., set gapWidth to 0%)
    elif style=="line":
        chart = LineChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        # Stacked bar plot
        chart.type = "line"

    # General Chart stuff
    chart.title = title
    chart.y_axis.title = ylabel
    chart.x_axis.title = xlabel
    chart.width = 30
    chart.height = 15

    # Add the chart to the sheet
    sheet.add_chart(chart, "D4")  # Adjust the position as needed

    # Save the workbook
    wb.save(filepath)


class cExcelFcts:
    '''
    This class is thightly coulpled with 2 excel templates. Originally designed for a specific use case and predefined system,
    class now accepts labels for components, that will be acessed in the visualization.
    If the used flixResults deffer from the original structure, these values can be adjusted
    '''
    def __init__(self,
                 calc: flixPostXL,
                 costs_label: str = "costs",
                 funding_label: str = "funding",
                 co2_label_fw: str = "CO2FW",
                 label_main_bus: str = "Fernwaerme",
                 label_demand: str = "Waermebedarf",
                 label_demand_losses="Netzverluste",
                 group_label_heat_demand_w_loss: str = "Wärmelast",
                 label_electr_production: str = "StromEinspeisung"
                 ):
        self.calc = calc

        self.costs_label = costs_label
        self.funding_label = funding_label
        self.co2_label_fw = co2_label_fw
        self.label_main_bus = label_main_bus
        self.label_demand = label_demand
        self.label_demand_losses = label_demand_losses
        self.group_label_heat_demand_w_loss = group_label_heat_demand_w_loss
        self.label_electr_production = label_electr_production

    def run_excel_graphics_years(self,
                                 short_version=False,
                                 custom_output_file_path: str = "default"):
        """
        Generate detailed annual comparison plots and save them to individual Excel workbooks for each year.

        Parameters:
        - calc (flixPostXL): Solved calculation of type flixPostXL.
        - short_version (bool): If True, generate a shortened version of the plots. Default is False.
        - custom_output_file_path (str): A custom output folder path for the Excel workbooks. default: Same as calc.

        Returns:
        None

        The function generates detailed annual comparison plots based on the provided calculation results (flixPostXL)
        and saves them to individual Excel workbooks. The workbooks are created using a template specified in the calculation.

        Plots and corresponding data are organized into different sheets within each workbook for the specified years:
        - "Wärmeerzeugung": Fernwärme generation data.
        - "Installierte Leistung": Installed capacity data.
        - "Wärmevollkosten": Heat full costs data.
        - "Wärmekosten Variabel": Variable heat costs data.
        - "Emissionen": Emissions data.
        - "Energieträger": Energy carriers data.
        - "Stromerzeugung": Electricity generation data.
        - "Speicherkapazität": Storage capacity data.
        - "Speicherfüllstand D": Daily storage fill level data.
        - "SpeicherFlows D": Daily storage flows data.
        - "WärmeErz-Last-D": Duration curve for heat generation and electricity prices, sorted by heat demand (Daily mean values).
        - "WärmeErz-Strom-D": Duration curves for heat generation and electricity prices, sorted by electricity prices (Daily mean values).

        If short_version is False (default), additional sheets are generated:
        - "WärmeErz-Last": Duration curve for heat generation and electricity prices, sorted by heat demand (Hourly values).
        - "WärmeErz-Strom": Duration curves for heat generation and electricity prices, sorted by electricity prices (Hourly values).
        - "Wärmeerzeugung_Februar": Hourly data for heat generation in February.
        - "Wärmeerzeugung_Juli": Hourly data for heat generation in July.
        - "WärmeErz-Last-DL-H": Annual load duration curves for heat generation. (Individually sorted for every generator)
        - "Speicher Summen": Hourly storage fill level data (Allocated over all storages).
        - "Speicherfüllstand H": Hourly storage fill level data for each individual storage.

        The Excel workbooks are saved in the specified output folder with filenames like
        "Jahr_{year}-{calc.infos['calculation']['name']}.xlsx".

        Example:
        ```
        calc = flixPostXL(...)  # Create or obtain flixPostXL instance
        run_excel_graphics_years(calc)  # Save the detailed workbooks in the default location
        run_excel_graphics_years(calc, short_version=True)  # Save shortened version of the workbooks in the default location
        run_excel_graphics_years(calc, custom_output_file_path="path/to/save/folder")  # Save the detailed workbooks in a custom location
        ```

        """
        if custom_output_file_path == "default":
            output_file_path = self.calc.folder
        else:
            output_file_path = custom_output_file_path

        print("Annual Plots to Excel...")

        # computation for the whole calculation

        df_fernwaerme_erz_nach_techn_D = self.get_fernwaerme_erz(resamply_by="d", rs_method="mean")  # Wärmeerzeugung

        df_installierte_leistung_Y = self.get_installierte_leistung(resamply_by="YE", rs_method="mean", flows=True,
                                                                     storage_capacity=False,
                                                                     grouped=True, actual_storage_capacity=False)

        df_waermekosten_vollkosten_D = self.get_waermekosten(with_fix_costs=True, resamply_by="d")

        df_waermekosten_varCosts_D = self.get_waermekosten(with_fix_costs=False, resamply_by="d")

        df_emissions_D = self.get_emissions(resamply_by="d", rs_method="sum")

        df_eingesetzte_energietraeger_D = self.get_eingesetzte_energietraeger(resamply_by="d", rs_method="mean")

        df_stromerzeugung_D = self.get_stromerzeugung(resamply_by="d")

        df_speicher_kapazitaet_D = self.get_speicher_kapazitaet(resamply_by="d",
                                                                 grouped=True, actual_storage_capacity=True)

        df_speicher_fuellstand_D = self.get_speicher_fuellstand("d", "mean", allocated=False)

        df_speicher_flows_D = self.get_speicher_flows("d", "mean", allocated=False)

        print("......computation of data for short version finished")
        if not short_version:
            df_fernwaerme_erz_nach_techn_H = self.get_fernwaerme_erz(resamply_by="h", rs_method="mean")
            df_speicher_fuellstand_H = self.get_speicher_fuellstand("h", "mean", allocated=False)
            df_speicher_fuellstand_H_alloc = self.get_speicher_fuellstand("h", "mean", allocated=True)

        # TODO: weitere Grafiken

        print("......computation of data finished")

        templ_path_excel_year = Path(__file__).resolve().parent / "resources" / "Template_Evaluation_Year.xlsx"

        for index, year in enumerate(self.calc.years):
            wb = load_workbook(templ_path_excel_year)
            filename = f"{self.calc.infos['calculation']['name']}__Jahr_{year}.xlsx"
            path_excel_year = os.path.join(output_file_path, filename)
            wb.save(path_excel_year)

            with pd.ExcelWriter(path_excel_year, mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:
                # Wärmeerzeugung nach Technologie
                df = df_fernwaerme_erz_nach_techn_D[df_fernwaerme_erz_nach_techn_D.index.year == year]
                df.to_excel(writer, index=True, sheet_name="Wärmeerzeugung")

                # Installierte Leistung nach Technologie
                df = df_installierte_leistung_Y[df_installierte_leistung_Y.index == year]
                df.to_excel(writer, index=True, sheet_name="Installierte Leistung")

                # Wärmevollkosten
                df = df_waermekosten_vollkosten_D[df_waermekosten_vollkosten_D.index.year == year]
                df.to_excel(writer, index=True, sheet_name="Wärmevollkosten")

                # Wärmekosten Betrieb
                df = df_waermekosten_varCosts_D[df_waermekosten_varCosts_D.index.year == year]
                df.to_excel(writer, index=True, sheet_name="Wärmekosten Variabel")

                # Emissionen
                df = df_emissions_D[df_emissions_D.index.year == year]
                df.to_excel(writer, index=True, sheet_name="Emissionen")

                # Energieträger
                df = df_eingesetzte_energietraeger_D[df_eingesetzte_energietraeger_D.index.year == year]
                df.to_excel(writer, index=True, sheet_name="Energieträger")

                # Stromerzeugung
                df = df_stromerzeugung_D[df_stromerzeugung_D.index.year == year]
                df.to_excel(writer, index=True, sheet_name="Stromerzeugung")

                # Speicherkapazität allokiert
                df = df_speicher_kapazitaet_D[df_speicher_kapazitaet_D.index.year == year]
                df.to_excel(writer, index=True, sheet_name="Speicherkapazität")

                # Speicherfüllstand nicht allokiert (Tageswerte)
                df = df_speicher_fuellstand_D[df_speicher_fuellstand_D.index.year == year]
                df.to_excel(writer, index=True, sheet_name="Speicherfüllstand D")

                # Speicherflows nicht allokiert (Tageswerte)
                df = df_speicher_flows_D[df_speicher_flows_D.index.year == year]
                df.to_excel(writer, index=True, sheet_name="SpeicherFlows D")

                # Wärmeerzeugung als Jahresdauerlinien (Tagesmittelwerte)
                df = df_fernwaerme_erz_nach_techn_D[df_fernwaerme_erz_nach_techn_D.index.year == year]
                df.sort_values("Wärmelast", ascending=False, ignore_index=True).to_excel(writer, index=True,
                                                                                         sheet_name="WärmeErz-Last-D")
                df.sort_values("Strompreis", ascending=False, ignore_index=True).to_excel(writer, index=True,
                                                                                          sheet_name="WärmeErz-Strom-D")

                print(f"......Year-{year} finished (short version)")
                if not short_version:
                    # Wärmeerzeugung als Jahresdauerlinien (Stundenwerte)
                    df = df_fernwaerme_erz_nach_techn_H[df_fernwaerme_erz_nach_techn_H.index.year == year]
                    df.sort_values("Wärmelast", ascending=False, ignore_index=True).to_excel(writer, index=True,
                                                                                             sheet_name="WärmeErz-Last")
                    df.sort_values("Strompreis", ascending=False, ignore_index=True).to_excel(writer, index=True,
                                                                                              sheet_name="WärmeErz-Strom")

                    # Wärmeerzeugung im Februar und Juli (Stundenwerte)
                    df = df_fernwaerme_erz_nach_techn_H[df_fernwaerme_erz_nach_techn_H.index.year == year]
                    df.loc[df.index.month == 2].to_excel(writer, index=True, sheet_name="Wärmeerzeugung_Februar")
                    df.loc[df.index.month == 7].to_excel(writer, index=True, sheet_name="Wärmeerzeugung_Juli")

                    # Jahresdauerlinien der einzelnen Wärmeerzeuger (Stundenwerte)
                    df = df_fernwaerme_erz_nach_techn_H[df_fernwaerme_erz_nach_techn_H.index.year == year]
                    df = pd.DataFrame(-np.sort(-df.values, axis=0), columns=df.columns)

                    df.to_excel(writer, index=True, sheet_name="WärmeErz-Last-DL-H")

                    # Speicherfüllstand (Stundenwerte) allokiert
                    df = df_speicher_fuellstand_H_alloc[df_speicher_fuellstand_H_alloc.index.year == year]
                    df.to_excel(writer, index=True, sheet_name="Speicher Summen")

                    # Speicherfüllstand (Stundenwerte) nicht allokiert
                    df = df_speicher_fuellstand_H[df_speicher_fuellstand_H.index.year == year]
                    df.to_excel(writer, index=True, sheet_name="Speicherfüllstand H")
                print(f"...Year-{year} finished")

                # TODO: weitere Grafiken

        print("...Annual Plots to Excel finished")

    def run_excel_graphics_main(self,
                                custom_output_file_path: str = "default"):
        """
        Generate annual comparison plots and save them to an Excel workbook.

        Parameters:
        - calc (flixPostXL): Solved calculation of type flixPostXL.
        - custom_output_file_path (str): A custom output file path (full path) for the Excel workbook. Default: Same as calc.

        Returns:
        None

        The function generates various annual comparison plots based on the provided calculation results (flixPostXL)
        and saves them to an Excel workbook. The workbook is created using a template specified in the calculation.

        Plots and corresponding data are organized into different sheets within the workbook:
        - "Waermelast und Verluste": Fernwärme load and losses data.
        - "Kostenübersicht": Costs overview data.
        - "Wärmeerzeugung": Fernwärme generation data.
        - "Installierte Leistung": Installed capacity data.
        - "Wärmevollkosten": Heat full costs data.
        - "Wärmekosten Variabel": Variable heat costs data.
        - "Emissionen": Emissions data.
        - "Energieträger": Energy carriers data.
        - "Stromerzeugung": Electricity generation data.
        - "Speicherkapazität": Storage capacity data.
        - "Speicher Summen": Summed storage fill level data.

        The Excel workbook is saved in the same folder as the calculation results with the filename
        "Jahresübersicht-{calc.infos['calculation']['name']}.xlsx". If a custom_output_file_path is provided,
        the workbook is saved at that location.

        Example:
        ```
        calc = flixPostXL(...)  # Create or obtain flixPostXL instance
        run_excel_graphics_main(calc)  # Save the workbook in the default location
        run_excel_graphics_main(calc, custom_output_file_path="path/to/save/file.xlsx")  # Save the workbook in a custom location
        ```

        """
        print("Overview Plots to Excel...")

        if custom_output_file_path == "default":
            output_file_path = self.calc.folder
        else:
            output_file_path = custom_output_file_path

        templ_path_excel_main = Path(__file__).resolve().parent / "resources" / "Template_Evaluation_Overview.xlsx"

        wb = load_workbook(templ_path_excel_main)
        filename = f"{self.calc.infos['calculation']['name']}__Jahresübersicht.xlsx"
        path_excel_main = os.path.join(output_file_path, filename)
        wb.save(path_excel_main)

        with pd.ExcelWriter(path_excel_main, mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:
            df = self.get_fernwaerme_last_and_loss("YE", "sum")
            df.to_excel(writer, index=True, sheet_name="Waermelast und Verluste")

            df = self.get_costs_and_funding("YE")
            df.to_excel(writer, index=True, sheet_name="Kostenübersicht")

            df = self.get_fernwaerme_erz("YE", "sum") / 1000
            df.to_excel(writer, index=True, sheet_name="Wärmeerzeugung")

            df = self.get_installierte_leistung(resamply_by="YE", rs_method="mean", flows=True, storage_capacity=False,
                                                 grouped=True, actual_storage_capacity=False)
            df.to_excel(writer, index=True, sheet_name="Installierte Leistung")

            df = self.get_waermekosten(with_fix_costs=True, resamply_by="YE")
            df.to_excel(writer, index=True, sheet_name="Wärmevollkosten")

            df = self.get_waermekosten(with_fix_costs=False, resamply_by="YE")
            df.to_excel(writer, index=True, sheet_name="Wärmekosten Variabel")

            df = self.get_emissions(resamply_by="YE", rs_method="sum")
            df.to_excel(writer, index=True, sheet_name="Emissionen")

            df = self.get_eingesetzte_energietraeger(resamply_by="YE", rs_method="sum") / 1000
            df.to_excel(writer, index=True, sheet_name="Energieträger")

            df = self.get_stromerzeugung(resamply_by="YE")
            df.to_excel(writer, index=True, sheet_name="Stromerzeugung")

            df_speicher_kapazitaet_Y = self.get_speicher_kapazitaet(resamply_by="YE",
                                                                     grouped=True, actual_storage_capacity=False)
            df_speicher_kapazitaet_Y.to_excel(writer, index=True, sheet_name="Speicherkapazität")

            df_speicher_fuellstand_sum_H = self.get_speicher_fuellstand("h", "mean", allocated=True).reset_index(
                drop=True)
            df_speicher_fuellstand_sum_H.to_excel(writer, index=True, sheet_name="Speicher Summen")

        print("...Overview Plots to Excel finished")

    def get_costs_and_funding(self, resamply_by):
        funding_var = self.calc.get_effect_results(self.funding_label, origin="operation", as_TS=True)
        funding_fix = self.calc.get_effect_results(self.funding_label, origin="invest", as_TS=True)
        costs_var = self.calc.get_effect_results(self.costs_label, origin="operation", as_TS=True)
        costs_fix = self.calc.get_effect_results(self.costs_label, origin="invest", as_TS=True)

        df = pd.DataFrame(data={"Fixkosten": costs_fix,
                                "Variable Kosten": costs_var,
                                "Förderung Invest": -funding_fix,
                                "Förderung Betrieb": -funding_var},
                          index=self.calc.timeSeries,
                          )
        df = resample_data(df, self.calc.years, resamply_by, "sum")
        return df

    def get_fernwaerme_erz(self, resamply_by, rs_method):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling

            if "d", Strompreis and Wärmelast are added to the DataFrame in first and second column
            if "YE", Wärmelast and sorages are not included
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value

        Returns
        -------
        pd.DataFrame
        '''

        if resamply_by == "YE":
            df_fernwaerme = self.calc.to_dataFrame(self.label_main_bus, "inout", grouped=False)  # ohne Wärmelast, ohne Speicher
            df_fernwaerme_grouped = self.calc.group_df_by_mapping(df_fernwaerme)
            df_fernwaerme_grouped.drop(columns=[self.group_label_heat_demand_w_loss], inplace=True)
        else:
            df_fernwaerme_grouped = self.calc.to_dataFrame(self.label_main_bus, "inout", grouped=True)
            df_fernwaerme_grouped[self.group_label_heat_demand_w_loss] = -1 * df_fernwaerme_grouped[self.group_label_heat_demand_w_loss]  # reinverting
            try:
                df_fernwaerme_grouped = pd.concat([df_fernwaerme_grouped, self.calc.getFuelCosts()["Strompreis"]], axis=1)
            except KeyError:
                print("Strompreis was not found and therefore can not be plotted")

        df_fernwaerme_erz_nach_techn = resample_data(df_fernwaerme_grouped, self.calc.years, resamply_by, rs_method)

        df_fernwaerme_erz_nach_techn = self.merge_into_dispatch_structure(df_fernwaerme_erz_nach_techn)

        return df_fernwaerme_erz_nach_techn

    def get_installierte_leistung(self, resamply_by, rs_method, flows: bool, storage_capacity: bool, grouped=False,
                                  actual_storage_capacity: bool = False):
        '''
        Parameters
        ----------
        calc1 : flix_results
            Calculation Object
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling

        Returns
        -------
        pd.DataFrame
        '''
        df_invest = self.calc.get_invest_results_as_TS(flows=flows, storages=storage_capacity,
                                                       grouped=grouped, actual_storage_capacity=actual_storage_capacity)
        df_invest = reorder_columns(df_invest)

        if df_invest.empty:
            return df_invest
        else:
            df_invest = resample_data(df_invest, self.calc.years, resamply_by, rs_method)
            df_invest = self.merge_into_dispatch_structure(df_invest)
            return df_invest

    def get_waermekosten(self, with_fix_costs, resamply_by):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value

        Returns
        -------
        pd.DataFrame
        '''
        heat = self.calc.to_dataFrame(self.label_demand, "in")

        if with_fix_costs:
            costs_total = pd.Series(self.calc.get_effect_results(effect_name=self.costs_label, origin="all", as_TS=True),
                                    index=self.calc.timeSeries)
        else:
            costs_total = pd.Series(self.calc.get_effect_results(effect_name=self.costs_label, origin="operation", as_TS=True),
                                    index=self.calc.timeSeries)

        # Unterschiedung zwischen Resampling
        if resamply_by == "d":
            rs_method_base = "h"
            new_columns = ["Tagesmittel", "Minimum (Stunde)", "Maximum (Stunde)"]
        elif resamply_by == "YE":
            rs_method_base = "d"
            new_columns = ["Jahresmittel", "Minimum (Tagesmittel)", "Maximum (Tagesmittel)"]
        else:
            raise ValueError(f"not implemented for resamply_by parameter: '{resamply_by}'")

        mean_costs_increment = resample_data(costs_total, self.calc.years, rs_method_base, "mean").iloc[:, 0]
        mean_heat_increment = resample_data(heat, self.calc.years, rs_method_base, "mean").iloc[:, 0]
        mean_costs_per_heat_increment = pd.DataFrame(mean_costs_increment / mean_heat_increment,
                                                     columns=["EURvarPerMWh"])

        minY = resample_data(mean_costs_per_heat_increment, self.calc.years, resamply_by, "min", rs_method_base)
        maxY = resample_data(mean_costs_per_heat_increment, self.calc.years, resamply_by, "max", rs_method_base)
        increment_sum_of_costs_total = resample_data(mean_costs_increment, self.calc.years, resamply_by, "sum",
                                                     rs_method_base).iloc[:, 0]
        increment_sum_of_heat_total = resample_data(mean_heat_increment, self.calc.years, resamply_by, "sum", rs_method_base).iloc[:,
                                      0]
        meanY = increment_sum_of_costs_total / increment_sum_of_heat_total

        df = pd.concat([meanY, minY, maxY], axis=1)
        df.columns = new_columns

        return df

    def get_emissions(self, resamply_by, rs_method):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value

        Returns
        -------
        pd.DataFrame
        '''
        heat = self.calc.to_dataFrame(self.label_demand, "in")

        CO2 = pd.DataFrame(self.calc.get_effect_results(effect_name=self.co2_label_fw, origin="operation", as_TS=True),
                           index=self.calc.timeSeries)

        CO2_per_increment = resample_data(CO2, self.calc.years, resamply_by, rs_method).iloc[:, 0]
        heat_per_increment = resample_data(heat, self.calc.years, resamply_by, rs_method).iloc[:, 0]
        CO2_per_heat = CO2_per_increment / heat_per_increment * 1000  # from t/MWh to kg/MWh
        df_emissions = pd.concat([CO2_per_heat.round(1), CO2_per_increment, heat_per_increment], axis=1)
        df_emissions.columns = ["kgCO2PerMWh", "tCO2absolut", "MWhabsolut"]

        return df_emissions

    def get_eingesetzte_energietraeger(self, resamply_by, rs_method):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value

        Returns
        -------
        pd.DataFrame
        '''
        df_sources = self.calc.get_sources_and_sinks(sources=True, sinks=False, sinks_n_sources=False)
        df = resample_data(df_sources, self.calc.years, resamply_by, rs_method)
        df = reorder_columns(df)

        return df

    def get_stromerzeugung(self, resamply_by):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value

        Returns
        -------
        resampled DataFrame with new columns:
            if resamply_by = "d": ["Tagesmittel", "Minimum (Stunde)", "Maximum (Stunde)"]
            if resamply_by = "YE": ["Jahresmittel", "Minimum (Tagesmittel)", "Maximum (Tagesmittel)"],
        '''
        df_stromerzeugung = self.calc.to_dataFrame(self.label_electr_production, "out",invert_Output=False)
        df = rs_in_two_steps(df_stromerzeugung, self.calc.years, resamply_by, "h")

        return df

    def get_speicher_kapazitaet(self, resamply_by, grouped, actual_storage_capacity: bool):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling

        Returns
        -------
        resampled DataFrame with capacity of all Storages
        '''
        invest_results_speicher = self.calc.get_invest_results_as_TS(storages=True, flows=False,
                                                                     actual_storage_capacity=actual_storage_capacity)
        if invest_results_speicher.empty:
            invest_results_speicher = pd.DataFrame(np.zeros(len(self.calc.timeSeries)), index=self.calc.timeSeries)
            invest_results_speicher.rename(columns={invest_results_speicher.columns[0]: "Speicher"}, inplace=True)

        elif grouped:
            invest_results_speicher = self.calc.group_df_by_mapping(invest_results_speicher)

        df = resample_data(invest_results_speicher, self.calc.years, resamply_by, "max")

        return df

    def get_speicher_fuellstand(self, resamply_by, rs_method, allocated):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value
        allocated : boolean
            True: sum of all storages in column one and netto flow of all Storages in column 2
            False: charge state of storage separately, no flow values

        Returns
        -------
        resampled DataFrame with total charge_state of all Storages
        '''
        df_speicher_chargeState = pd.DataFrame(index=self.calc.timeSeries)
        df_speicher_nettoFlow = pd.DataFrame(index=self.calc.timeSeries)

        list_of_speicher = [comp.label for comp in self.calc.comp_posts if comp.type == "cStorage"]

        for comp in list_of_speicher:
            df_speicher_chargeState[comp] = self.calc.results[comp]["charge_state"][:-1]  # without the last step
            df_speicher_nettoFlow[comp] = self.calc.results[comp]["nettoFlow"]

        if allocated:
            charge_state_sum = df_speicher_chargeState.sum(axis=1)
            netto_flow_sum = df_speicher_nettoFlow.sum(axis=1)*-1

            df = pd.concat([charge_state_sum, netto_flow_sum], axis=1)
            df.columns = ["Gesamtspeicherstand", "Nettospeicherflow"]
            df = resample_data(df, self.calc.years, resamply_by, rs_method)
        else:
            df = resample_data(df_speicher_chargeState, self.calc.years, resamply_by, rs_method)

        return df

    def get_fernwaerme_last_and_loss(self, resamply_by, rs_method):
        df_demand = self.calc.to_dataFrame(self.label_demand, "in")
        df_loss = self.calc.to_dataFrame(self.label_demand_losses, "in")
        df = pd.concat([df_demand, df_loss], axis=1)
        df_summed = resample_data(df, self.calc.years, resamply_by, rs_method)
        df_verluste_summed = (df_summed.iloc[:, 1] / df_summed.sum(axis=1) * 100).rename("Verlust[%]").round(2)

        return pd.concat([df_summed, df_verluste_summed], axis=1)

    def get_speicher_flows(self, resamply_by, rs_method, allocated):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value
        allocated : boolean
            True: sum of all storages in column one and netto flow of all Storages in column 2
            False: charge state of storage separately, no flow values

        Returns
        -------
        resampled DataFrame with total charge_state of all Storages
        '''
        df_speicher_nettoFlow = pd.DataFrame(index=self.calc.timeSeries)

        list_of_speicher = [comp.label for comp in self.calc.comp_posts if comp.type == "cStorage"]

        for comp in list_of_speicher:
            df_speicher_nettoFlow[comp] = self.calc.results[comp]["nettoFlow"]*-1

        if allocated:
            df = df_speicher_nettoFlow.sum(axis=1)
            df = resample_data(df, self.calc.years, resamply_by, rs_method)
            df.columns = ["Nettospeicherflow"]
        else:
            df = resample_data(df_speicher_nettoFlow, self.calc.years, resamply_by, rs_method)

        return df

    def merge_into_dispatch_structure(self, df:pd.DataFrame) -> pd.DataFrame:
        '''
        Brings a dataframe into a predefined structure for dispatch evaluation.
        Has space for 9 undefined columns
        '''
        # Step 1: Create an empty DataFrame with specific column names
        fixed_columns_1 = ['TAB', 'Geothermie', 'Abwärme', 'WP', 'WP_2', 'EHK', 'KWK_Gas', 'KWK_H2',
                           'Kessel_Gas', 'Kessel_H2', 'Speicher_S', 'Speicher_L', 'Kühler']  # First 11 fixed columns
        undefined_columns = ['U1', 'U2', 'U3', 'U4', 'U5', 'U6', 'U7', 'U8', 'U9']  # 8 undefined placeholders
        fixed_columns_2 = ['others', 'Wärmelast', 'Strompreis']  # Last 2 fixed columns

        # Combine all parts into the final column structure
        all_columns = fixed_columns_1 + undefined_columns + fixed_columns_2

        # Step 2: Create the target DataFrame with this structure, initially filled with None
        df_target = pd.DataFrame(columns=all_columns, index=df.index)

        # String formattin to prevent unintended behaviour
        df.columns = (df.columns
                      .str.replace('ae', 'ä')
                      .str.replace('oe', 'ö')
                      .str.replace('ue', 'ü')
                      .str.strip()
                      )
        df.columns = [col[0].upper() + col[1:] for col in df.columns]

        # Merge logic
        # Directly assign matched columns
        for col in df.columns.intersection(df_target.columns):
            df_target[col] = df[col]

        # Handle unmatched columns by placing them into the undefined placeholders
        unmatched_columns = df.columns.difference(df_target.columns)
        unmatched_columns = sorted(unmatched_columns, key=lambda x: x.lower())  # sorting alphabetically
        for i, col in enumerate(unmatched_columns):
            if i < len(undefined_columns):  # Ensure there's an available placeholder
                df_target[undefined_columns[i]] = df[col]
                df_target = df_target.rename(columns={undefined_columns[i]: col})
            else:
                df_target['others'] = df[[col for col in unmatched_columns[i:]]].sum(axis=1)

        # removing all values when all nan values
        nan_columns = df_target.columns[df_target.isnull().all()]
        rename_dict = {col: "."*i for i, col in enumerate(nan_columns)}
        df_target = df_target.rename(columns=rename_dict)
        return df_target


def write_bus_results_to_excel(calc:flixPostXL, resample_by: Literal["YE", "d", "h"] = "d",
                               custom_output_file_path: str = "default"):
    """
    Save the in- and out-flows of every bus to an Excel file.

    Parameters
    ----------
    calc : flixPostXL
        The flixPostXL object containing the calculation results.
    resample_by : str, optional
        The time frequency for resampling data (e.g., 'd' for daily), by default "d".
        Allowed values are 'YE' (yearly), 'd' (daily), and 'h' (hourly).
    custom_output_file_path : str, optional
        Custom path to save the Excel file

    Returns
    -------
    None
    """
    print(f"...Writing Bus Results ({resample_by}) to Excel...")

    if custom_output_file_path == "default":
        output_file_path = calc.folder
    else:
        output_file_path = custom_output_file_path

    filename = f"{calc.infos['calculation']['name']}__Buses_{resample_by}.xlsx"
    path_excel = os.path.join(output_file_path, filename)

    for bus_name in calc.buses:
        data = calc.to_dataFrame(busOrComp=bus_name, direction="inout", invert_Output=True) * -1
        data = resample_data(data_frame=data, target_years=calc.years, resampling_by=resample_by,
                             resampling_method="sum")
        df_to_excel_w_chart(data, path_excel, bus_name, "MWh", "Time")

    print(f"......Buses ({resample_by}) finished")

def write_component_results_to_excel(calc: flixPostXL, resample_by: Literal["YE", "d", "h"] = "d",
                                     custom_output_file_path: str = "default"):
    """
    Save the in- and out-flows of every component to an Excel file.

    Parameters
    ----------
    calc : flixPostXL
        The flixPostXL object containing the calculation results.
    resample_by : str, optional
        The time frequency for resampling data (e.g., 'd' for daily), by default "d".
        Allowed values are 'YE' (yearly), 'd' (daily), and 'h' (hourly).
    custom_output_file_path : str, optional
        Custom path to save the Excel file

    Returns
    -------
    None
    """
    print(f"...Writing Components Results ({resample_by}) to Excel...")

    if custom_output_file_path == "default":
        output_file_path = calc.folder
    else:
        output_file_path = custom_output_file_path

    filename = f"{calc.infos['calculation']['name']}__Comps_{resample_by}.xlsx"
    path_excel = os.path.join(output_file_path, filename)

    for comp_name in calc.comps:
        data = calc.to_dataFrame(busOrComp=comp_name, direction="inout", invert_Output=True) * -1
        data = resample_data(data_frame=data, target_years=calc.years, resampling_by=resample_by,
                             resampling_method="sum")
        df_to_excel_w_chart(data, path_excel, comp_name, "MWh", "Time")

    print(f"......Components ({resample_by}) finished")

def write_effect_results_to_excel_shares(calc, custom_output_file_path: str = "default"):
    """
    Save summarized effects data to an Excel file.

    Parameters
    ----------
    calc : flixPostXL
        The flixPostXL object containing the calculation results.
    Returns
    -------
    None
    """
    print(f"...Writing Effects Results (Shares) to Excel...")

    if custom_output_file_path == "default":
        output_file_path = calc.folder
    else:
        output_file_path = custom_output_file_path

    filename = f"{calc.infos['calculation']['name']}__Effects_Shares.xlsx"
    path_excel = os.path.join(output_file_path, filename)

    df_effects_sum = pd.DataFrame()
    for effect_name, effect in calc.results["globalComp"].items():
        if effect_name == "penalty":
            continue
        new_df = pd.DataFrame(calc.get_effect_results(effect_name=effect_name, origin="all", as_TS=False,
                                                              shares=True), index = [effect_name]).T
        df_effects_sum = pd.concat([df_effects_sum, new_df], axis=0)
    df_to_excel_w_chart(df_effects_sum, path_excel, "Effects_SUM_Shares", "See Legend", "Component", style="bar")

    df_effects_op = pd.DataFrame()
    for effect_name, effect in calc.results["globalComp"].items():
        if effect_name == "penalty":
            continue
        new_df = pd.DataFrame(calc.get_effect_results(effect_name=effect_name, origin="operation", as_TS=False,
                                                      shares=True), index=[effect_name]).T
        df_effects_op = pd.concat([df_effects_op, new_df], axis=0)
    df_to_excel_w_chart(df_effects_op, path_excel, "Effects_OP_Shares", "diverse", "Component", style="bar")

    df_effects_inv = pd.DataFrame()
    for effect_name, effect in calc.results["globalComp"].items():
        if effect_name == "penalty":
            continue
        new_df = pd.DataFrame(calc.get_effect_results(effect_name=effect_name, origin="invest", as_TS=False,
                                                      shares=True), index=[effect_name]).T
        df_effects_inv = pd.concat([df_effects_inv, new_df], axis=0)
    df_to_excel_w_chart(df_effects_inv, path_excel, "Effects_Inv_Shares", "diverse", "Component", style="bar")

    print(f"......Effects (Shares) finished")

def write_effect_results_per_comp_to_excel(calc, custom_output_file_path: str = "default"):
    """
    Save summarized effects data to an Excel file.

    Parameters
    ----------
    calc : flixPostXL
        The flixPostXL object containing the calculation results.
    Returns
    -------
    None
    """
    print(f"...Writing Effects Results (Shares) to Excel...")

    if custom_output_file_path == "default":
        output_file_path = calc.folder
    else:
        output_file_path = custom_output_file_path

    filename = f"{calc.infos['calculation']['name']}__Effects_per_comp.xlsx"
    path_excel = os.path.join(output_file_path, filename)

    results_operation = {}
    results_invest = {}
    results_total = {}
    for effect in calc.results["globalComp"].keys():
        if effect == "penalty":
            continue
        results_operation[effect] = {}
        results_invest[effect] = {}
        results_total[effect] = {}
        for comp in calc.comps:
            operation_costs = calc.get_effect_result_of_comp_without_shares(effect, comp, "operation")
            invest_costs = calc.get_effect_result_of_comp_without_shares(effect, comp, "invest")
            if operation_costs:
                results_operation[effect][comp] = operation_costs
            if invest_costs:
                results_invest[effect][comp] = invest_costs
            if operation_costs or invest_costs:
                results_total[effect][comp] = operation_costs + invest_costs

    results_operation["heat"] = {}
    results_invest["heat"] = {}
    results_total["heat"] = {}

    for comp in calc.comps:
        heat = 0
        for flow in calc.flows:
            if flow.from_node == comp and flow.to_node == "Fernwaerme":
                heat += sum(flow.results["val"])
            elif flow.from_node == "Fernwaerme" and flow.to_node == comp:
                heat -= sum(flow.results["val"])
        if heat:
            results_operation["heat"][comp] = heat
            results_invest["heat"][comp] = heat
            results_total["heat"][comp] = heat


    df_effects_invest = pd.DataFrame(results_invest)
    df_effects_operation = pd.DataFrame(results_operation)
    df_effects_total = pd.DataFrame(results_total)

    try:
        df_effects_invest["costs per heat [€/MWh]"] = df_effects_invest["costs"]/df_effects_invest["heat"]
        df_effects_invest["costs per heat (incl. funding) [€/MWh]"] = (df_effects_invest["costs"] - df_effects_invest["funding"] ) / df_effects_invest["heat"]

        df_effects_operation["costs per heat [€/MWh]"] = df_effects_operation["costs"]/df_effects_operation["heat"]
        df_effects_operation["costs per heat (incl. funding) [€/MWh]"] = (df_effects_operation["costs"] - df_effects_operation["funding"] ) / df_effects_operation["heat"]

        df_effects_total["costs per heat [€/MWh]"] = df_effects_total["costs"]/df_effects_total["heat"]
        df_effects_total["costs per heat (incl. funding) [€/MWh]"] = (df_effects_total["costs"] - df_effects_total["funding"] ) / df_effects_total["heat"]
    except KeyError as e:
        print(f"Caluclation of costs per heat and cost per heat (including funding) Threw an exception: {e}")

    df_to_excel_w_chart(df_effects_invest, path_excel, "Effects_Per_Comp_Invest", "See Legend", "Component Name", style="bar", bar_style="clustered")
    df_to_excel_w_chart(df_effects_operation, path_excel, "Effects_Per_Comp_Operation", "See Legend", "Component Name",
                        style="bar", bar_style="clustered")
    df_to_excel_w_chart(df_effects_total, path_excel, "Effects_Per_Comp_Total", "See Legend", "Component Name",
                        style="bar", bar_style="clustered")


    print(f"......Effects (Shares) finished")

def write_effect_results_to_excel(calc, resample_by: Literal["YE", "d", "h"] = "d",
                                  custom_output_file_path: str = "default"):
    """
    Save summarized effects data to an Excel file.

    Parameters
    ----------
    calc : flixPostXL
        The flixPostXL object containing the calculation results.
    resample_by : str, optional
        The time frequency for resampling data (e.g., 'd' for daily), by default "d".
        Allowed values are 'YE' (yearly), 'd' (daily), and 'h' (hourly).
    custom_output_file_path : str, optional
        Custom path to save the Excel file

    Returns
    -------
    None
    """
    print(f"...Writing Effects Results ({resample_by}) to Excel...")

    if custom_output_file_path == "default":
        output_file_path = calc.folder
    else:
        output_file_path = custom_output_file_path

    filename = f"{calc.infos['calculation']['name']}__Effects-{resample_by}.xlsx"
    path_excel = os.path.join(output_file_path, filename)

    df_effects_sum = pd.DataFrame()
    for effect_name, effect in calc.results["globalComp"].items():
        if effect_name == "penalty":
            continue
        df_effects_sum[effect_name] = calc.get_effect_results(effect_name=effect_name, origin="all", as_TS=True,
                                                              shares=False)
    df_effects_sum = resample_data(data_frame=df_effects_sum, target_years=calc.years, resampling_by=resample_by,
                                   resampling_method="sum")
    df_to_excel_w_chart(df_effects_sum, path_excel, "Effects_SUM", "See Legend", "Time", style="line")

    df_effects_op = pd.DataFrame()
    for effect_name, effect in calc.results["globalComp"].items():
        if effect_name == "penalty":
            continue
        df_effects_op[effect_name] = calc.get_effect_results(effect_name=effect_name, origin="operation",
                                                             as_TS=True, shares=False)
    df_effects_op = resample_data(data_frame=df_effects_op, target_years=calc.years, resampling_by=resample_by,
                                  resampling_method="sum")
    df_to_excel_w_chart(df_effects_op, path_excel, "Effects_OP", "diverse", "Time", style="line")

    df_effects_inv = pd.DataFrame()
    for effect_name, effect in calc.results["globalComp"].items():
        if effect_name == "penalty":
            continue
        df_effects_inv[effect_name] = calc.get_effect_results(effect_name=effect_name, origin="invest", as_TS=True,
                                                              shares=False)
    df_effects_inv = resample_data(data_frame=df_effects_inv, target_years=calc.years, resampling_by=resample_by,
                                   resampling_method="sum")
    df_to_excel_w_chart(df_effects_inv, path_excel, "Effects_Inv", "diverse", "Time", style="line")

    print(f"......Effects ({resample_by}) finished")

def visualize_results(calc_results: flixPostXL,
                      effect_shares: bool = True, effects_per_comp: bool = True,
                      buses_yearly: bool = True, comps_yearly: bool = True, effects_yearly: bool = True,
                      buses_daily: bool = True, comps_daily: bool = True, effects_daily: bool = True,
                      buses_hourly: bool = False, comps_hourly: bool = False,
                      effects_hourly: bool = False) -> None:
    """
    Visualizes the results of a flixPostXL object.

    * The overview results are mainly used to compare yearly mean values
      between different years.

    * The annual results are used to go into detail about the heating
      production and storage usage in each year.

    * The buses results are used to look at all uses of energy balance.

    * The comps results are used to look at all Transformation processes
      in the different components.

    * The effects results are used to look at all effects. Effects are
      Costs, CO2 Funding, etc.

    * Daily mean values are enough for most use cases.

    * Hourly values are good for in-depth examinations, but take a long
      time to extract and save.

    * TAKE CARE: Writing hourly data to excel takes a significant amount of time for
      big Models with many Components.

    Parameters:
        overview (bool): Whether to write overview graphics. Default is True.
        annual_results (bool): Whether to write annual results graphics. Default is True.
        effects_shares (bool): Whether to write effect shares to excel. Default is True.
        buses_yearly (bool): Whether to write annual results for buses to excel. Default is True.
        comps_yearly (bool): Whether to write annual results for components to excel. Default is True.
        effects_yearly (bool): Whether to write annual results for effects to excel. Default is True.
        buses_daily (bool): Whether to write daily results for buses to excel. Default is True.
        comps_daily (bool): Whether to write daily results for components to excel. Default is True.
        effects_daily (bool): Whether to write daily results for effects to excel. Default is True.
        buses_hourly (bool): Whether to write hourly results for buses to excel. Default is False.
        comps_hourly (bool): Whether to write hourly results for components to excel. Default is False.
        effects_hourly (bool): Whether to write hourly results for effects to excel. Default is False.

    Returns:
        flixPostXL: The calculated results.
    """
    calc_results.visual_representation(save_to=os.path.join(calc_results.folder, f"{calc_results.label}-Model_structure.html"))

    print("Writing Results to Excel (YE)...")
    if buses_yearly: write_bus_results_to_excel(calc_results, "YE")
    if effects_yearly: write_effect_results_to_excel(calc_results, "YE")
    if comps_yearly: write_component_results_to_excel(calc_results, "YE")
    if effect_shares: write_effect_results_to_excel_shares(calc_results)
    if effects_per_comp: write_effect_results_per_comp_to_excel(calc_results)
    print("...Results to Excel (YE) finished...")

    print("Writing Results to Excel (d)...")
    if buses_daily: write_bus_results_to_excel(calc_results, "d")
    if effects_daily: write_effect_results_to_excel(calc_results, "d")
    if comps_daily: write_component_results_to_excel(calc_results,  "d")
    print("...Results to Excel (d) finished...")

    print("Writing results to Excel (h)...")
    if buses_hourly: write_bus_results_to_excel(calc_results, "h")
    if effects_hourly: write_effect_results_to_excel(calc_results,  "h")
    if comps_hourly: write_component_results_to_excel(calc_results,  "h")
    print("...Results to Excel (h) finished...")


################## PDF - OUTPUT ####################
def create_report(calc: flixPostXL, path: str = 'report.pdf', connected_to: str = "Fernwaerme", chunk_size: int = 4):
    print(f"Creating Report of Components connected to '{connected_to}'")
    flows_to_plot = []
    for flow in calc.flows:
        if flow.to_node == connected_to or flow.from_node == connected_to:
            if np.sum(flow.results["val"]) >= 1:
                flows_to_plot.append(flow.label_full)
    flows_to_plot.sort()

    with PdfPages(path) as pdf:
        for chunk in [flows_to_plot[i:i + chunk_size] for i in range(0, len(flows_to_plot), chunk_size)]:
            fig, axes = plt.subplots(len(chunk), 1, figsize=(8.27, 11.69 / 4 * len(chunk)), sharex=True,
                                     sharey=True)  # A4 size
            if not isinstance(axes, np.ndarray):  # If only one item in batch, axes is not a list...
                axes = np.array([axes])

            for ax, flow_name in zip(axes.flatten(), chunk):
                fig, ax = calc.plotOperationColorMap(flow_name, nbPeriods=365 * len(calc.years), fig=fig, ax=ax,
                                                     ylabel="time of day", xlabel="day")
                ax.set_title(flow_name)  # Set individual title for each subplot

            pdf.savefig(fig)
            plt.close()

def create_report_grouped(calc: flixPostXL, path: str = 'report.pdf', connected_to: str = "Fernwaerme", chunk_size: int = 4) -> None:
    print(f"Creating Report of Components connected to '{connected_to}'")
    # Filtering and sorting
    flows_to_plot = {"others": []}
    for flow in calc.flows:
        if flow.to_node == connected_to or flow.from_node == connected_to:
            if np.sum(flow.results["val"]) >= 1:
                group = flow.group if flow.group else "others"
                if group in flows_to_plot.keys():
                    flows_to_plot[group].append(flow.label_full)
                else:
                    flows_to_plot[group] = [flow.label_full]
    for key in flows_to_plot:
        flows_to_plot[key].sort()
    groups = list(flows_to_plot.keys())
    groups.sort()


    with PdfPages(path) as pdf:
        for group in groups:
            items_in_group = len(flows_to_plot[group])
            for chunk in [flows_to_plot[group][i:i + chunk_size] for i in range (0, items_in_group, chunk_size)]:
                fig, axes = plt.subplots(len(chunk), 1, figsize=(8.27, 11.69/4 * len(chunk)), sharex=True, sharey=True)  #A4 size
                if not isinstance(axes, np.ndarray):  # If only one item in batch, axes is not a list...
                    axes = np.array([axes])

                for ax, flow_name in zip(axes.flatten(), chunk):
                    fig, ax = calc.plotOperationColorMap(flow_name, nbPeriods=365 * len(calc.years), fig=fig, ax=ax,
                                                         ylabel="time of day", xlabel="day")
                    ax.set_title(flow_name)  # Set individual title for each subplot

                fig.text(0, 1, f"Group: {group}", ha='left', va='center', fontsize=14, weight='bold')  # Add big header
                pdf.savefig(fig, bbox_inches='tight')
                plt.close()

def create_report_per_comp(calc: flixPostXL, path: str = 'report.pdf') -> None:
    # Filtering and sorting
    components = {}
    for flow in calc.flows:
        if flow.comp in components:
            components[flow.comp].append(flow.label_full)
        else:
            components[flow.comp] = [flow.label_full]

    for key in components:
        components[key].sort()
    components_sorted = list(components.keys())
    components_sorted.sort()


    with PdfPages(path) as pdf:
        for comp_name in components_sorted:
            nrOfSubplots = len(components[comp_name])
            fig, axes = plt.subplots(nrOfSubplots, 1, figsize=(8.27, 11.69/4 * nrOfSubplots) , sharex=True, sharey=True)  # A4 size
            if not isinstance(axes, np.ndarray):  # If only one item in batch, axes is not a list...
                axes = np.array([axes])
            for ax, flow_name in zip(axes.flatten(), components[comp_name]):
                fig, ax = calc.plotOperationColorMap(flow_name, nbPeriods=365 * len(calc.years), fig=fig, ax=ax,
                                                     ylabel="time of day", xlabel="day")
                ax.set_title(flow_name)  # Set individual title for each subplot
            fig.text(0.5, 0.95, f"{comp_name}", ha='center', va='center', fontsize=14, weight='bold')  # Add big header

            pdf.savefig(fig, bbox_inches='tight')
            plt.close()
